from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "SOC2 AWS Mapping"

# Define the table data with columns A-L as headers
headers = [
    "A: Control ID",
    "B: Control Title",
    "C: Full Description",
    "D: Primary AWS Service",
    "E: Supporting Services",
    "F: Implementation Steps",
    "G: Evidence Type",
    "H: Where to Find Evidence",
    "I: Difficulty",
    "J: Monthly Cost",
    "K: Setup Time",
    "L: Notes"
]

descriptions = [
    "CC6.1, CC6.2, etc.",
    'Short name (e.g., "MFA Required")',
    "What the control actually requires",
    "Main service that addresses this (IAM, Config, etc.)",
    "Other AWS services that help",
    "Brief how-to (detailed guide goes in separate docs)",
    "What auditors will want to see",
    "Specific AWS console location",
    "Easy / Medium / Hard",
    "Estimate in $",
    "Hours to implement",
    "Common pitfalls, tips, etc."
]

# Set column widths for all columns
for i in range(1, 13):
    ws.column_dimensions[get_column_letter(i)].width = 25

# Define styles
header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Write headers (row 1)
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = cell_alignment

# Write descriptions (row 2)
for col_num, description in enumerate(descriptions, 1):
    cell = ws.cell(row=2, column=col_num)
    cell.value = description
    cell.alignment = cell_alignment
    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

# Set row heights
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40

# Save the workbook
filename = "soc2-aws-mapping-template.xlsx"
wb.save(filename)
print(f"Excel spreadsheet '{filename}' created successfully!")
